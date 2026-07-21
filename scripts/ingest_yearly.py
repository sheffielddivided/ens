#!/usr/bin/env python3
"""Phase 2 -- parse the yearly SI-unit Excel file into ``data/yearly.json``.

The yearly workbook is the authoritative history (1972 -> latest). Its exact
sheet layout is not published, and it uses merged cells, footnotes and possibly
multi-level headers, so this parser is **inspection-first**: it discovers the
structure of every sheet at run time rather than hard-coding cell positions.

Two layouts are supported and auto-detected per sheet:

* **matrix** -- one measure per sheet (sheet name says oil / gas / water /
  injection / flare / fuel / export). Years run along one axis and field names
  along the other; orientation is detected by locating the run of year cells.
* **long/tabular** -- a single sheet with ``field``, ``year`` and one column
  per measure.

If a sheet matches neither and cannot be interpreted, it is skipped with a WARN;
if *no* sheet yields data, the run fails with ``SourceFormatError`` rather than
writing an empty/dubious file.

Use ``--inspect`` to dump sheet names, dimensions and a preview before trusting
the parse. Run ``python scripts/ingest_yearly.py --help`` for all options.
"""

from __future__ import annotations

import argparse
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import common as C  # noqa: E402

YEAR_MIN, YEAR_MAX = 1970, 2035

# Sheet-name keyword -> measure. Order matters: more specific first.
_MEASURE_RULES = [
    (("gas", "inject"), "gas_injection"),
    (("gas", "export"), "gas_export"),
    (("oil", "export"), "oil_export"),
    (("water", "inject"), "water_injection"),
    (("water", "prod"), "water"),
    (("oil", "prod"), "oil"),
    (("sales", "gas"), "gas"),
    (("gas", "sale"), "gas"),
    (("flar",), "flare"),
    (("fuel",), "fuel"),
    (("water",), "water"),
    (("oil",), "oil"),
    (("gas",), "gas"),
]
# Danish equivalents folded in.
_DANISH_SYNONYMS = {
    "olie": "oil", "vand": "water", "injektion": "inject",
    "fakling": "flare", "flaring": "flare", "braendsel": "fuel",
    "brndsel": "fuel", "produktion": "prod", "eksport": "export",
    "salgsgas": "sales gas",
}


def classify_measure(sheet_name: str) -> str | None:
    """Map a sheet name to a measure key, or None if it names no measure."""
    s = C.normalize_field(sheet_name).replace("_", " ")
    for da, en in _DANISH_SYNONYMS.items():
        if da in s:
            s += " " + en
    for keywords, measure in _MEASURE_RULES:
        if all(k in s for k in keywords):
            return measure
    return None


def _is_year(v) -> int | None:
    n = C.parse_number(v)
    if n is None:
        return None
    if n == int(n) and YEAR_MIN <= int(n) <= YEAR_MAX:
        return int(n)
    return None


def _load_grids(path: Path) -> list[tuple[str, list[list]]]:
    """Return [(sheet_name, grid)] with merged cells forward-filled."""
    import io
    import openpyxl

    # ENS serves the workbook via an extension-less /media/<id>/download URL,
    # so the cache file has no ".xlsx" suffix. Passing a file-like object makes
    # openpyxl validate by content instead of by filename extension.
    data = Path(path).read_bytes()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 -- turn into a clear source error
        head = data[:8]
        raise C.SourceFormatError(
            f"could not open yearly workbook as .xlsx (first bytes={head!r}); "
            f"a legacy .xls or an error page would land here: {exc}"
        ) from exc
    out = []
    for ws in wb.worksheets:
        grid = [[c.value for c in row] for row in ws.iter_rows()]
        if not grid:
            out.append((ws.title, grid))
            continue
        width = max(len(r) for r in grid)
        for r in grid:
            r.extend([None] * (width - len(r)))
        # Forward-fill each merged range with its top-left value.
        for rng in ws.merged_cells.ranges:
            tl = ws.cell(rng.min_row, rng.min_col).value
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    grid[rr - 1][cc - 1] = tl
        out.append((ws.title, grid))
    return out


def _find_year_axis(grid: list[list]):
    """Locate where year values live.

    Returns ("row", idx, {col: year}) if years span a header row, or
    ("col", idx, {row: year}) if they span a column, or None.
    """
    best_row = (0, -1, {})
    for ri, row in enumerate(grid):
        years = {ci: y for ci, v in enumerate(row) if (y := _is_year(v)) is not None}
        if len(years) > best_row[0]:
            best_row = (len(years), ri, years)
    best_col = (0, -1, {})
    ncols = max((len(r) for r in grid), default=0)
    for ci in range(ncols):
        years = {}
        for ri, row in enumerate(grid):
            if ci < len(row) and (y := _is_year(row[ci])) is not None:
                years[ri] = y
        if len(years) > best_col[0]:
            best_col = (len(years), ci, years)

    # Require at least 3 years on an axis to accept it as the year axis.
    if best_row[0] < 3 and best_col[0] < 3:
        return None
    if best_row[0] >= best_col[0]:
        return ("row", best_row[1], best_row[2])
    return ("col", best_col[1], best_col[2])


def _label_of(grid, ri, ci) -> str:
    try:
        v = grid[ri][ci]
    except IndexError:
        return ""
    return "" if v is None else str(v).strip()


def parse_matrix_sheet(
    grid: list[list], measure: str, raw_labels: dict | None = None
) -> dict[tuple[str, int], float]:
    """Parse a single-measure matrix sheet into {(field_slug, year): value}.

    When ``raw_labels`` (a defaultdict(Counter)) is supplied, the original text
    label seen for each slug is recorded so fields.json can carry aliases.
    """
    axis = _find_year_axis(grid)
    if axis is None:
        raise C.SourceFormatError(
            f"could not locate a year axis in a '{measure}' sheet"
        )
    kind, idx, years = axis
    out: dict[tuple[str, int], float] = {}

    def note(slug: str, label: str) -> None:
        if raw_labels is not None and label:
            raw_labels[slug][label] += 1

    if kind == "row":
        # Header row = years; a label column holds field names (left of years).
        year_cols = sorted(years)
        label_col = _pick_label_column(grid, before_col=min(year_cols))
        for ri, row in enumerate(grid):
            if ri == idx:
                continue
            label = _label_of(grid, ri, label_col)
            slug = C.normalize_field(label)
            if not slug or C.is_total_label(label):
                continue
            note(slug, label)
            for ci in year_cols:
                val = C.parse_number(row[ci] if ci < len(row) else None)
                if val is not None:
                    out[(slug, years[ci])] = val
    else:
        # Year column = years; a header row holds field names (above years).
        year_rows = sorted(years)
        label_row = _pick_label_row(grid, before_row=min(year_rows), year_col=idx)
        for ci in range(max(len(r) for r in grid)):
            if ci == idx:
                continue
            label = _label_of(grid, label_row, ci)
            slug = C.normalize_field(label)
            if not slug or C.is_total_label(label):
                continue
            note(slug, label)
            for ri in year_rows:
                row = grid[ri]
                val = C.parse_number(row[ci] if ci < len(row) else None)
                if val is not None:
                    out[(slug, years[ri])] = val
    return out


def _distinct_text(values) -> int:
    """Count distinct non-empty, non-year text values.

    A merged title row/column collapses to a single distinct value, whereas a
    real field-name header has many -- this is what separates them.
    """
    seen = {
        str(v).strip() for v in values
        if isinstance(v, str) and v.strip() and _is_year(v) is None
    }
    return len(seen)


def _pick_label_column(grid, before_col: int) -> int:
    """Choose the field-name column left of the years.

    Prefer the column with the most *distinct* text values; break ties toward
    the column nearest the year data (a merged title has 1 distinct value).
    """
    best = None  # (distinct, col_index) -- higher col wins ties (closer to data)
    for ci in range(before_col):
        distinct = _distinct_text(row[ci] for row in grid if ci < len(row))
        if distinct == 0:
            continue
        cand = (distinct, ci)
        if best is None or cand > best:
            best = cand
    return best[1] if best else 0


def _pick_label_row(grid, before_row: int, year_col: int) -> int:
    """Choose the field-name header row above the years.

    Prefer the row with the most *distinct* text values; break ties toward the
    row nearest the year data (a merged title collapses to 1 distinct value).
    """
    best = None  # (distinct, row_index) -- higher row wins ties (closer to data)
    for ri in range(before_row):
        row = grid[ri]
        distinct = _distinct_text(v for ci, v in enumerate(row) if ci != year_col)
        if distinct == 0:
            continue
        cand = (distinct, ri)
        if best is None or cand > best:
            best = cand
    return best[1] if best else max(0, before_row - 1)


# --------------------------------------------------------------------------- #
# Long / tabular sheet
# --------------------------------------------------------------------------- #
_HEADER_ALIASES = {
    "field": {"field", "felt", "field_name", "feltnavn"},
    "year": {"year", "aar", "ar"},
    "oil": {"oil", "olie", "oil_production"},
    "gas": {"gas", "sales_gas", "salgsgas", "gas_production"},
    "water": {"water", "vand", "water_production"},
    "gas_injection": {"gas_injection", "gas_inj", "injected_gas"},
    "water_injection": {"water_injection", "water_inj", "injected_water"},
    "flare": {"flare", "flaring", "fakling"},
    "fuel": {"fuel", "braendsel", "fuel_gas"},
    "gas_export": {"gas_export", "eksport_gas"},
    "oil_export": {"oil_export", "eksport_olie"},
}


def _match_header(cell: str) -> str | None:
    slug = C.normalize_field(cell)
    for canonical, aliases in _HEADER_ALIASES.items():
        if slug in aliases:
            return canonical
    return None


def try_parse_long_sheet(grid: list[list]) -> list[dict] | None:
    """Parse a long/tabular sheet if it has field+year+measure headers."""
    for ri, row in enumerate(grid):
        mapping = {}
        for ci, v in enumerate(row):
            if isinstance(v, str):
                canon = _match_header(v)
                if canon and canon not in mapping:
                    mapping[canon] = ci
        if "field" in mapping and "year" in mapping and (
            set(mapping) & set(C.CORE_MEASURES)
        ):
            return _read_long_rows(grid, ri, mapping)
    return None


def _read_long_rows(grid, header_ri, mapping) -> list[dict]:
    records = []
    for row in grid[header_ri + 1:]:
        label = _label_of_row(row, mapping["field"])
        slug = C.normalize_field(label)
        if not slug or C.is_total_label(label):
            continue
        year = _is_year(row[mapping["year"]] if mapping["year"] < len(row) else None)
        if year is None:
            continue
        rec = {"field": slug, "year": year, "_label": label}
        for measure, ci in mapping.items():
            if measure in ("field", "year"):
                continue
            val = C.parse_number(row[ci] if ci < len(row) else None)
            if val is not None:
                rec[measure] = val
        if any(m in rec for m in (*C.CORE_MEASURES, *C.OPTIONAL_MEASURES)):
            records.append(rec)
    return records


def _label_of_row(row, ci) -> str:
    v = row[ci] if ci < len(row) else None
    return "" if v is None else str(v).strip()


# --------------------------------------------------------------------------- #
# Stacked single-sheet layout (the real ENS yearly file)
# --------------------------------------------------------------------------- #
# The published yearly workbook is one sheet with measure blocks stacked
# vertically: a title row ("Oil, thousand cubic meters" / "Gas, million normal
# cubic meters" / "Water, ...") or a sub-label ("Export", "Fuel", "Flare",
# "Injection"), then a horizontal year header (1972 ...), then one row per field
# ending in "Total". Export/Injection belong to whichever section (oil/gas/
# water) is currently in effect.
def _count_year_cells(row) -> int:
    return sum(1 for v in row if _is_year(v) is not None)


def _first_text(row) -> str:
    for v in row:
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _classify_block(titles: list[str], section: str | None):
    """Return (measure, new_section) for a block given its title rows."""
    j = " ".join(t.lower() for t in titles)
    if "inject" in j:
        return ("water_injection" if section == "water" else "gas_injection"), section
    if "export" in j:
        return ("oil_export" if section == "oil" else "gas_export"), section
    if "fuel" in j:
        return "fuel", section
    if "flar" in j:
        return "flare", section
    if "oil" in j:
        return "oil", "oil"
    if "gas" in j:
        return "gas", "gas"
    if "water" in j:
        return "water", "water"
    return None, section


def try_parse_stacked_sheet(grid: list[list], raw_labels: dict | None = None):
    """Parse a single sheet of vertically-stacked measure blocks.

    Returns a list of {field, year, <measures>} records, or None if the sheet
    is not this shape (needs at least two year-header rows so the simpler
    one-measure-per-sheet layout is left to the matrix parser).
    """
    year_headers = [ri for ri, row in enumerate(grid) if _count_year_cells(row) >= 5]
    if len(year_headers) < 2:
        return None

    section = None
    agg: dict[tuple[str, int], dict] = {}
    for idx, h in enumerate(year_headers):
        prev = year_headers[idx - 1] if idx > 0 else -1
        # Gather the title rows directly above this year header (stop at a blank
        # row or the previous block).
        titles: list[str] = []
        r = h - 1
        while r > prev:
            txt = _first_text(grid[r])
            if not txt:
                break
            titles.insert(0, txt)
            r -= 1
        measure, section = _classify_block(titles, section)
        if measure is None:
            continue
        years = {ci: _is_year(v) for ci, v in enumerate(grid[h]) if _is_year(v) is not None}
        nxt = year_headers[idx + 1] if idx + 1 < len(year_headers) else len(grid)
        for ri in range(h + 1, nxt):
            row = grid[ri]
            label = row[0] if row else None
            slug = C.normalize_field(label)
            if not slug or C.is_total_label(label):
                continue
            got = False
            for ci, yr in years.items():
                val = C.parse_number(row[ci] if ci < len(row) else None)
                if val is not None:
                    agg.setdefault((slug, yr), {"field": slug, "year": yr})[measure] = val
                    got = True
            if got and raw_labels is not None:
                raw_labels[slug][str(label).strip()] += 1
    if not agg:
        return None
    return [agg[k] for k in sorted(agg)]


# --------------------------------------------------------------------------- #
# Workbook dispatcher
# --------------------------------------------------------------------------- #
def parse_workbook(path: Path) -> tuple[list[dict], dict]:
    """Return (records, report). Each record is one field/year with measures."""
    grids = _load_grids(path)
    by_key: dict[tuple[str, int], dict] = {}
    report = {"sheets": [], "measures_found": Counter(), "raw_labels": defaultdict(Counter)}

    for name, grid in grids:
        if not grid or not any(any(c is not None for c in r) for r in grid):
            report["sheets"].append({"sheet": name, "result": "empty"})
            continue

        # 1) Try long/tabular first (self-describing headers).
        long_records = try_parse_long_sheet(grid)
        if long_records:
            for rec in long_records:
                key = (rec["field"], rec["year"])
                by_key.setdefault(key, {"field": rec["field"], "year": rec["year"]})
                if rec.get("_label"):
                    report["raw_labels"][rec["field"]][rec["_label"]] += 1
                for m, v in rec.items():
                    if m in ("field", "year", "_label"):
                        continue
                    by_key[key][m] = v
                    report["measures_found"][m] += 1
            report["sheets"].append(
                {"sheet": name, "result": "long", "records": len(long_records)}
            )
            continue

        # 2) Stacked single-sheet layout (the real ENS yearly workbook).
        stacked = try_parse_stacked_sheet(grid, report["raw_labels"])
        if stacked:
            for rec in stacked:
                key = (rec["field"], rec["year"])
                by_key.setdefault(key, {"field": rec["field"], "year": rec["year"]})
                for m, v in rec.items():
                    if m in ("field", "year"):
                        continue
                    by_key[key][m] = v
                    report["measures_found"][m] += 1
            report["sheets"].append(
                {"sheet": name, "result": "stacked", "records": len(stacked)}
            )
            continue

        # 3) Otherwise treat as a single-measure matrix.
        measure = classify_measure(name)
        if measure is None:
            report["sheets"].append({"sheet": name, "result": "skipped-unknown-measure"})
            C.warn(f"sheet '{name}': no measure keyword and not tabular; skipped")
            continue
        try:
            cells = parse_matrix_sheet(grid, measure, report["raw_labels"])
        except C.SourceFormatError as exc:
            report["sheets"].append({"sheet": name, "result": f"error: {exc}"})
            C.warn(f"sheet '{name}': {exc}")
            continue
        for (slug, year), val in cells.items():
            key = (slug, year)
            by_key.setdefault(key, {"field": slug, "year": year})
            by_key[key][measure] = val
            report["measures_found"][measure] += 1
        report["sheets"].append(
            {"sheet": name, "result": f"matrix:{measure}", "values": len(cells)}
        )

    records = [by_key[k] for k in sorted(by_key)]
    # Drop records with no core measure at all (noise rows).
    records = [r for r in records if any(m in r for m in C.CORE_MEASURES)]
    if not records:
        raise C.SourceFormatError(
            "yearly workbook produced no records -- structure not recognised. "
            "Re-run with --inspect and update the parser."
        )
    return records, report


# --------------------------------------------------------------------------- #
# fields.json
# --------------------------------------------------------------------------- #
_KNOWN_OPERATORS = {
    # Best-effort seed; unknown fields stay null. Operators change over time.
    "dan": "TotalEnergies", "gorm": "TotalEnergies", "halfdan": "TotalEnergies",
    "tyra": "TotalEnergies", "skjold": "TotalEnergies", "valdemar": "TotalEnergies",
    "roar": "TotalEnergies", "svend": "TotalEnergies", "harald": "TotalEnergies",
    "siri": "INEOS", "nini": "INEOS", "cecilie": "INEOS",
    "south_arne": "Hess", "syd_arne": "Hess",
}


def build_fields(records: list[dict], raw_labels: dict[str, Counter]) -> dict:
    seen: dict[str, dict] = {}
    for r in records:
        f = r["field"]
        info = seen.setdefault(
            f, {"slug": f, "display_name": None, "aliases": set(),
                "first_year": r["year"], "last_year": r["year"], "operator": None}
        )
        info["first_year"] = min(info["first_year"], r["year"])
        info["last_year"] = max(info["last_year"], r["year"])
    fields = {}
    for slug, info in sorted(seen.items()):
        aliases = raw_labels.get(slug)
        if aliases:
            display = aliases.most_common(1)[0][0]
            alias_list = sorted(aliases)
        else:
            display = slug.replace("_", " ").title()
            alias_list = []
        fields[slug] = {
            "slug": slug,
            "display_name": display,
            "aliases": alias_list,
            "first_year": info["first_year"],
            "last_year": info["last_year"],
            "operator": _KNOWN_OPERATORS.get(slug),
        }
    return {
        "schema_version": C.SCHEMA_VERSION,
        "generated_at": C.utc_now_iso(),
        "fields": fields,
    }


# --------------------------------------------------------------------------- #
# Spot check
# --------------------------------------------------------------------------- #
def _spot_check(records: list[dict]) -> None:
    """Log a few aggregate figures so a human can sanity-check the parse."""
    by_year_oil: dict[int, float] = defaultdict(float)
    for r in records:
        if "oil" in r:
            by_year_oil[r["year"]] += r["oil"]
    if by_year_oil:
        C.info("spot check -- total oil per year (sum over fields, first/last 3):")
        yrs = sorted(by_year_oil)
        sample = yrs[:3] + (["..."] if len(yrs) > 6 else []) + yrs[-3:]
        for y in sample:
            if y == "...":
                C.info("     ...")
            else:
                C.info(f"     {y}: {by_year_oil[y]:,.1f} ({C.UNIT_DEFINITIONS['oil']})")


# --------------------------------------------------------------------------- #
# Orchestration
# --------------------------------------------------------------------------- #
def inspect(path: Path) -> None:
    grids = _load_grids(path)
    C.info(f"workbook {path.name}: {len(grids)} sheet(s)")
    for name, grid in grids:
        nrows = len(grid)
        ncols = max((len(r) for r in grid), default=0)
        measure = classify_measure(name)
        C.info(f"  sheet '{name}': {nrows}x{ncols}  measure={measure}")
        for r in grid[:6]:
            preview = [("" if c is None else str(c))[:14] for c in r[:8]]
            C.info("      | " + " | ".join(preview))


def ingest(
    *, source_url: str | None, offline: bool, force: bool,
    xlsx_path: Path | None, do_inspect: bool,
) -> int:
    fetcher = C.Fetcher(offline=offline)
    if xlsx_path is None:
        if source_url is None:
            idx = C.read_json(C.INDEX_PATH, default={}) or {}
            y = idx.get("yearly")
            if not y or not y.get("url"):
                C.error("no yearly URL: run build_index.py first or pass --url/--xlsx")
                return 1
            source_url = y["url"]
        filename = C.safe_filename(source_url)
        fetcher.fetch(source_url, filename, force=force, binary=True)
        xlsx_path = fetcher.cache_path(filename)
    else:
        source_url = source_url or f"file://{xlsx_path}"

    if do_inspect:
        inspect(xlsx_path)
        return 0

    C.info(f"parsing yearly workbook: {xlsx_path}")
    records, report = parse_workbook(xlsx_path)

    retrieved = C.utc_now_iso()
    for r in records:
        r["preliminary"] = False
        r["source_url"] = source_url
        r["retrieved_at"] = retrieved
    records = C.sort_records(records, monthly=False)

    yearly = {
        "schema_version": C.SCHEMA_VERSION,
        "unit_definitions": C.UNIT_DEFINITIONS,
        "generated_at": retrieved,
        "source_url": source_url,
        "records": records,
    }
    C.write_json_stable(C.YEARLY_PATH, yearly, volatile_keys=("generated_at",))
    fields = build_fields(records, report["raw_labels"])
    C.write_json_stable(C.FIELDS_PATH, fields, volatile_keys=("generated_at",))

    C.info(f"wrote {len(records)} yearly records over "
           f"{len(fields['fields'])} fields to {C.YEARLY_PATH}")
    C.info(f"measures found: {dict(report['measures_found'])}")
    _spot_check(records)
    return 0


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Parse the yearly SI-unit Excel into data/yearly.json + fields.json",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--url", help="Yearly Excel URL (defaults to index.json's yearly).")
    p.add_argument("--xlsx", type=Path, help="Use a local .xlsx instead of downloading.")
    p.add_argument("--offline", action="store_true", help="Use only cached downloads.")
    p.add_argument("--force", action="store_true", help="Re-download even if cached.")
    p.add_argument("--inspect", action="store_true",
                   help="Print sheet structure and exit (no writing).")
    args = p.parse_args(argv)
    try:
        return ingest(source_url=args.url, offline=args.offline, force=args.force,
                      xlsx_path=args.xlsx, do_inspect=args.inspect)
    except Exception as exc:  # noqa: BLE001
        C.error(f"ingest_yearly failed: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
