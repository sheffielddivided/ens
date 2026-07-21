#!/usr/bin/env python3
"""Generate tests/fixtures/yearly.xlsx -- a SYNTHETIC yearly workbook.

It mimics the *structure* of the ENS yearly SI-unit file (one sheet per measure,
merged title rows, a Total row/column, both axis orientations) so the parser is
exercised without a real download. Numbers are made up. Regenerate with:

    python tests/fixtures/make_xlsx_fixture.py
"""
from pathlib import Path

import openpyxl

OUT = Path(__file__).resolve().parent / "yearly.xlsx"

YEARS = [2018, 2019, 2020, 2021, 2022, 2023]
FIELDS = ["Dan", "Gorm", "Halfdan"]

# measure -> {field -> base value}
DATA = {
    "oil": {"Dan": 300, "Gorm": 150, "Halfdan": 900},
    "gas": {"Dan": 200, "Gorm": 90, "Halfdan": 400},
    "water": {"Dan": 1200, "Gorm": 600, "Halfdan": 2500},
}


def _series(base: float, year: int, i: int) -> float:
    # Deterministic gentle variation so values differ per year/field.
    return round(base * (1 + 0.03 * (year - 2018)) + i * 1.5, 1)


def main() -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: "Oil production" -- years vertical (col A), fields across header.
    ws = wb.create_sheet("Oil production")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Oil production in the Danish North Sea (1000 m3)"
    ws.append([])  # blank spacer row
    ws.append(["Year"] + FIELDS + ["Total"])
    for y in YEARS:
        vals = [_series(DATA["oil"][f], y, i) for i, f in enumerate(FIELDS)]
        ws.append([y] + vals + [round(sum(vals), 1)])

    # Sheet 2: "Sales gas production" -- years horizontal, fields vertical.
    ws = wb.create_sheet("Sales gas production")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Sales gas production (mio. Nm3)"
    ws.append([])
    ws.append(["Field"] + YEARS)
    for i, f in enumerate(FIELDS):
        ws.append([f] + [_series(DATA["gas"][f], y, i) for y in YEARS])
    ws.append(["Total"] + [
        round(sum(_series(DATA["gas"][f], y, i) for i, f in enumerate(FIELDS)), 1)
        for y in YEARS
    ])

    # Sheet 3: "Water production" -- years vertical, with a footnote marker.
    ws = wb.create_sheet("Water production")
    ws.merge_cells("A1:E1")
    ws["A1"] = "Water production (1000 m3)"
    ws.append([])
    ws.append(["Year", "Dan", "Gorm", "Halfdan*", "Total"])
    for y in YEARS:
        vals = [_series(DATA["water"][f], y, i) for i, f in enumerate(FIELDS)]
        ws.append([y] + vals + [round(sum(vals), 1)])

    # Sheet 4: an injection sheet (optional measure) -- years vertical.
    ws = wb.create_sheet("Water injection")
    ws.append(["Year", "Dan", "Gorm", "Halfdan"])
    for y in YEARS:
        ws.append([y, _series(500, y, 0), _series(200, y, 1), _series(800, y, 2)])

    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
